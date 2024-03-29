import os
import openpyxl
from products.models import Product
from company.models import Company
from products.serializers import ProductSerializer

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

product_data = dict()


def change_image_path(image):
    filename = os.path.split(image)[1]
    dir_path = os.path.split(image)[1]
    return dir_path, filename


def upload_products(request):
    user_email = request.user.email
    user_resources = os.path.join(BASE_DIR, 'static/media/{}/resources/'.format(user_email))
    files = os.listdir(user_resources)
    products_db = os.path.join(user_resources, 'products.xlsx')
    for file in files:
        if str(file) != 'products.xlsx':
            upload_xl = os.path.join(user_resources, file)
            return upload_xl, products_db


def delete_uploaded_file(request):
    uploaded_xl = upload_products(request)[0]
    os.remove(uploaded_xl)


def spreadsheet_db(request):
    upload_wb = openpyxl.load_workbook(upload_products(request)[0], data_only=True)
    products_wb = openpyxl.load_workbook(upload_products(request)[1], data_only=True)
    print(upload_wb)

    upload_sheet_1 = upload_wb['Sheet1']
    products_sheet_1 = products_wb['Sheet1']

    up_max_row = upload_sheet_1.max_row
    up_max_col = upload_sheet_1.max_column

    prod_max_row = products_sheet_1.max_row

    wanted_column = ['product-name', 'product-category', 'product-description', 'product-url', 'product-image']

    upload_wanted = dict()

    for wanted in wanted_column:
        for column in range(1, up_max_col + 1):
            column_name = upload_sheet_1.cell(row=1, column=column).value
            if column_name and wanted == column_name.lower():
                upload_wanted[f'{wanted}'] = column   # upload_wanted['product-name'] = 1   - example

    count = 0

    company = Company.objects.get(account=request.user.account)

    for row in range(2, up_max_row + 1):
        for column_name, column_number in upload_wanted.items():
            if column_name == wanted_column[0] and upload_sheet_1.cell(row, column_number).value:
                products_sheet_1.cell(prod_max_row + row - 1, 1).value = \
                    upload_sheet_1.cell(row, column_number).value
                product_data['product_name'] = upload_sheet_1.cell(row, column_number).value
            if column_name == wanted_column[1] and upload_sheet_1.cell(row, column_number).value:
                products_sheet_1.cell(prod_max_row + row - 1, 2).value = \
                    upload_sheet_1.cell(row, column_number).value
                product_data['category'] = upload_sheet_1.cell(row, column_number).value
            if column_name == wanted_column[2] and upload_sheet_1.cell(row, column_number).value:
                products_sheet_1.cell(prod_max_row + row - 1, 3).value = \
                    upload_sheet_1.cell(row, column_number).value
                product_data['description'] = upload_sheet_1.cell(row, column_number).value
            if column_name == wanted_column[3] and upload_sheet_1.cell(row, column_number).value:
                products_sheet_1.cell(prod_max_row + row - 1, 4).value = \
                    upload_sheet_1.cell(row, column_number).value
                product_data['url'] = upload_sheet_1.cell(row, column_number).value
            if column_name == wanted_column[4] and upload_sheet_1.cell(row, column_number).value:
                products_sheet_1.cell(prod_max_row + row - 1, 5).value = \
                    upload_sheet_1.cell(row, column_number).value
                product_data['image'] = upload_sheet_1.cell(row, column_number).value
            product_data['company'] = company

        print(product_data)

        # product_serializer = ProductSerializer(data=product_data)

        Product.objects.create(**product_data)

        count += 1

    products_wb.save(upload_products(request)[1])
    delete_uploaded_file(request)

    return count


def write_product_from_db_spreadsheet(request, product_instance):
    user_email = request.user.email
    product_db = os.path.join(BASE_DIR, 'static/media/{}/resources/products.xlsx'.format(user_email))
    product_name = product_instance.product_name
    category = product_instance.category
    description = product_instance.description
    url = product_instance.url
    image = product_instance.image.path

    products_wb = openpyxl.load_workbook(product_db)

    products_sheet_1 = products_wb['Sheet1']
    prod_max_row = products_sheet_1.max_row

    products_sheet_1.cell(prod_max_row + 1, 1).value = product_name
    products_sheet_1.cell(prod_max_row + 1, 2).value = category
    products_sheet_1.cell(prod_max_row + 1, 3).value = description
    products_sheet_1.cell(prod_max_row + 1, 4).value = url
    products_sheet_1.cell(prod_max_row + 1, 5).value = image

    products_wb.save(product_db)
